import os
import csv
from openpyxl import Workbook
from openpyxl import load_workbook


from keras.models import Sequential
from keras.layers import Dense
from keras.models import Sequential
from keras.layers import Dense
from keras.utils import to_categorical
import numpy

# get current working directory and append Data folder
TEAMS = []
path = os.getcwd() + '\Data'
os.chdir(path)
directory = os.listdir(path)
wb = load_workbook('AdvancedStats.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
Matches = []
i = 0
home = -1
X = []
Y = []
for x in range(1, 5128, 4):
    if sheet.cell(row=x, column=1).value not in TEAMS:      # home team
        TEAMS.append(sheet.cell(row=x, column=1).value)
    if sheet.cell(row=x, column=2).value not in TEAMS:      # away team
        TEAMS.append(sheet.cell(row=x, column=2).value)
    if float(sheet.cell(row=x+2, column=14).value) > float(sheet.cell(row=x+2, column=15).value):
        # winner = sheet.cell(row=x, column=1).value
        home = 1
    else:
        # winner = sheet.cell(row=x, column=2).value
        home = 0
    Y.append(home)
    i = 2
    temp = []
    while i < 13:
        temp.append(float(sheet.cell(row=x+2, column=i).value))
        temp.append(float(sheet.cell(row=x+3, column=i).value))
        i += 1
    X.append(temp)

    Matches.append([sheet.cell(row=x, column=1).value, sheet.cell(row=x, column=2).value, home])
# at this point TEAMS contains each team that participated in last season
# matches should now have the following structure:
# Matches[i] = [team1, team2, homeTeamWin], homeTeamWin is either 0 or 1
print(Matches)
# create model

model = Sequential()
'''
model.add(Dense(50,input_shape=(22,), activation='relu'))
model.add(Dense(40, activation='relu'))
model.add(Dense(30, activation='relu'))
model.add(Dense(50, activation='relu'))
model.add(Dense(35, activation='relu'))
model.add(Dense(15, activation='relu'))
model.add(Dense(5, activation='relu'))
'''

'''
# WORKING MODEL
model.add(Dense(30,input_shape=(22,), activation='relu'))
model.add(Dense(5, activation='relu'))
model.add(Dense(1, activation='sigmoid'))
'''
model.add(Dense(1,input_shape=(22,), activation='relu'))
model.compile(loss='binary_crossentropy', optimizer='adam', metrics=['accuracy'])
model.fit(X, Y, epochs=250, batch_size=10)
scores = model.evaluate(X, Y)
print("\n%s: %.2f%%" % (model.metrics_names[1], scores[1]*100))

# *********************************************************************************************************************
# ACTUAL PREDICTION CODE BELOW ****************************************************************************************
# *********************************************************************************************************************

wb = load_workbook('StatsSoFar.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
Matches = []
i = 0
home = -1
Xhome = []
Xaway = []
homeTeam = "Charlotte Hornets"
awayTeam = "Orlando Magic"
for x in range(1, 1284, 4):
    if sheet.cell(row=x, column=1).value == homeTeam:      # home team
        i = 2
        temp = []
        while i < 13:
            temp.append(float(sheet.cell(row=x + 2, column=i).value))
            i += 1
        Xhome.append(temp)
    if sheet.cell(row=x, column=2).value == homeTeam:      # the home team today was away then
        i = 2
        temp = []
        while i < 13:
            temp.append(float(sheet.cell(row=x + 3, column=i).value))
            i += 1
        Xhome.append(temp)
    if sheet.cell(row=x, column=1).value == awayTeam:      # the away team today was home team
        i = 2
        temp = []
        while i < 13:
            temp.append(float(sheet.cell(row=x + 2, column=i).value))
            i += 1
        Xaway.append(temp)
    if sheet.cell(row=x, column=2).value == awayTeam:      # away team
        i = 2
        temp = []
        while i < 13:
            temp.append(float(sheet.cell(row=x + 3, column=i).value))
            i += 1
        Xaway.append(temp)


avgHome = [float(sum(col))/len(col) for col in zip(*Xhome)]
avgAway = [float(sum(col))/len(col) for col in zip(*Xaway)]
finalPredictionInput = avgHome + avgAway
numpy.expand_dims(finalPredictionInput, axis=0)
predictions = model.predict(numpy.array([finalPredictionInput, ]))
notRounded = [x[0] for x in predictions]
rounded = [round(x[0]) for x in predictions]
# print(Y)
print(homeTeam, awayTeam)
print(rounded, notRounded)
Xhome = []
Xaway = []
homeTeam = "Indiana Pacers"
awayTeam = "New York Knicks"
for x in range(1, 1284, 4):
    if sheet.cell(row=x, column=1).value == homeTeam:      # home team
        i = 2
        temp = []
        while i < 13:
            temp.append(float(sheet.cell(row=x + 2, column=i).value))
            i += 1
        Xhome.append(temp)
    if sheet.cell(row=x, column=2).value == homeTeam:      # the home team today was away then
        i = 2
        temp = []
        while i < 13:
            temp.append(float(sheet.cell(row=x + 3, column=i).value))
            i += 1
        Xhome.append(temp)
    if sheet.cell(row=x, column=1).value == awayTeam:      # the away team today was home team
        i = 2
        temp = []
        while i < 13:
            temp.append(float(sheet.cell(row=x + 2, column=i).value))
            i += 1
        Xaway.append(temp)
    if sheet.cell(row=x, column=2).value == awayTeam:      # away team
        i = 2
        temp = []
        while i < 13:
            temp.append(float(sheet.cell(row=x + 3, column=i).value))
            i += 1
        Xaway.append(temp)


avgHome = [float(sum(col))/len(col) for col in zip(*Xhome)]
avgAway = [float(sum(col))/len(col) for col in zip(*Xaway)]
finalPredictionInput = avgHome + avgAway
numpy.expand_dims(finalPredictionInput, axis=0)
predictions = model.predict(numpy.array([finalPredictionInput, ]))
notRounded = [x[0] for x in predictions]
rounded = [round(x[0]) for x in predictions]
# print(Y)
print(homeTeam, awayTeam)
print(rounded, notRounded)
Xhome = []
Xaway = []
homeTeam = "Philadelphia 76ers"
awayTeam = "Phoenix Suns"
for x in range(1, 1284, 4):
    if sheet.cell(row=x, column=1).value == homeTeam:      # home team
        i = 2
        temp = []
        while i < 13:
            temp.append(float(sheet.cell(row=x + 2, column=i).value))
            i += 1
        Xhome.append(temp)
    if sheet.cell(row=x, column=2).value == homeTeam:      # the home team today was away then
        i = 2
        temp = []
        while i < 13:
            temp.append(float(sheet.cell(row=x + 3, column=i).value))
            i += 1
        Xhome.append(temp)
    if sheet.cell(row=x, column=1).value == awayTeam:      # the away team today was home team
        i = 2
        temp = []
        while i < 13:
            temp.append(float(sheet.cell(row=x + 2, column=i).value))
            i += 1
        Xaway.append(temp)
    if sheet.cell(row=x, column=2).value == awayTeam:      # away team
        i = 2
        temp = []
        while i < 13:
            temp.append(float(sheet.cell(row=x + 3, column=i).value))
            i += 1
        Xaway.append(temp)


avgHome = [float(sum(col))/len(col) for col in zip(*Xhome)]
avgAway = [float(sum(col))/len(col) for col in zip(*Xaway)]
finalPredictionInput = avgHome + avgAway
numpy.expand_dims(finalPredictionInput, axis=0)
predictions = model.predict(numpy.array([finalPredictionInput, ]))
notRounded = [x[0] for x in predictions]
rounded = [round(x[0]) for x in predictions]
# print(Y)
print(homeTeam, awayTeam)

print(rounded, notRounded)
Xhome = []
Xaway = []
homeTeam = "Atlanta Hawks"
awayTeam = "Brooklyn Nets"
for x in range(1, 1284, 4):
    if sheet.cell(row=x, column=1).value == homeTeam:      # home team
        i = 2
        temp = []
        while i < 13:
            temp.append(float(sheet.cell(row=x + 2, column=i).value))
            i += 1
        Xhome.append(temp)
    if sheet.cell(row=x, column=2).value == homeTeam:      # the home team today was away then
        i = 2
        temp = []
        while i < 13:
            temp.append(float(sheet.cell(row=x + 3, column=i).value))
            i += 1
        Xhome.append(temp)
    if sheet.cell(row=x, column=1).value == awayTeam:      # the away team today was home team
        i = 2
        temp = []
        while i < 13:
            temp.append(float(sheet.cell(row=x + 2, column=i).value))
            i += 1
        Xaway.append(temp)
    if sheet.cell(row=x, column=2).value == awayTeam:      # away team
        i = 2
        temp = []
        while i < 13:
            temp.append(float(sheet.cell(row=x + 3, column=i).value))
            i += 1
        Xaway.append(temp)


avgHome = [float(sum(col))/len(col) for col in zip(*Xhome)]
avgAway = [float(sum(col))/len(col) for col in zip(*Xaway)]
finalPredictionInput = avgHome + avgAway
numpy.expand_dims(finalPredictionInput, axis=0)
predictions = model.predict(numpy.array([finalPredictionInput, ]))
notRounded = [x[0] for x in predictions]
rounded = [round(x[0]) for x in predictions]
# print(Y)
print(homeTeam, awayTeam)

print(rounded, notRounded)

Xhome = []
Xaway = []
homeTeam = "Boston Celtics"
awayTeam = "Milwaukee Bucks"
for x in range(1, 1284, 4):
    if sheet.cell(row=x, column=1).value == homeTeam:      # home team
        i = 2
        temp = []
        while i < 13:
            temp.append(float(sheet.cell(row=x + 2, column=i).value))
            i += 1
        Xhome.append(temp)
    if sheet.cell(row=x, column=2).value == homeTeam:      # the home team today was away then
        i = 2
        temp = []
        while i < 13:
            temp.append(float(sheet.cell(row=x + 3, column=i).value))
            i += 1
        Xhome.append(temp)
    if sheet.cell(row=x, column=1).value == awayTeam:      # the away team today was home team
        i = 2
        temp = []
        while i < 13:
            temp.append(float(sheet.cell(row=x + 2, column=i).value))
            i += 1
        Xaway.append(temp)
    if sheet.cell(row=x, column=2).value == awayTeam:      # away team
        i = 2
        temp = []
        while i < 13:
            temp.append(float(sheet.cell(row=x + 3, column=i).value))
            i += 1
        Xaway.append(temp)


avgHome = [float(sum(col))/len(col) for col in zip(*Xhome)]
avgAway = [float(sum(col))/len(col) for col in zip(*Xaway)]
finalPredictionInput = avgHome + avgAway
numpy.expand_dims(finalPredictionInput, axis=0)
predictions = model.predict(numpy.array([finalPredictionInput, ]))
notRounded = [x[0] for x in predictions]
rounded = [round(x[0]) for x in predictions]
# print(Y)
print(homeTeam, awayTeam)
print(rounded, notRounded)


Xhome = []
Xaway = []
homeTeam = "Chicago Bulls"
awayTeam = "Cleveland Cavaliers"
for x in range(1, 1284, 4):
    if sheet.cell(row=x, column=1).value == homeTeam:      # home team
        i = 2
        temp = []
        while i < 13:
            temp.append(float(sheet.cell(row=x + 2, column=i).value))
            i += 1
        Xhome.append(temp)
    if sheet.cell(row=x, column=2).value == homeTeam:      # the home team today was away then
        i = 2
        temp = []
        while i < 13:
            temp.append(float(sheet.cell(row=x + 3, column=i).value))
            i += 1
        Xhome.append(temp)
    if sheet.cell(row=x, column=1).value == awayTeam:      # the away team today was home team
        i = 2
        temp = []
        while i < 13:
            temp.append(float(sheet.cell(row=x + 2, column=i).value))
            i += 1
        Xaway.append(temp)
    if sheet.cell(row=x, column=2).value == awayTeam:      # away team
        i = 2
        temp = []
        while i < 13:
            temp.append(float(sheet.cell(row=x + 3, column=i).value))
            i += 1
        Xaway.append(temp)


avgHome = [float(sum(col))/len(col) for col in zip(*Xhome)]
avgAway = [float(sum(col))/len(col) for col in zip(*Xaway)]
finalPredictionInput = avgHome + avgAway
numpy.expand_dims(finalPredictionInput, axis=0)
predictions = model.predict(numpy.array([finalPredictionInput, ]))
notRounded = [x[0] for x in predictions]
rounded = [round(x[0]) for x in predictions]
# print(Y)
print(homeTeam, awayTeam)
print(rounded, notRounded)


Xhome = []
Xaway = []
homeTeam = "Memphis Grizzlies"
awayTeam = "Minnesota Timberwolves"
for x in range(1, 1284, 4):
    if sheet.cell(row=x, column=1).value == homeTeam:      # home team
        i = 2
        temp = []
        while i < 13:
            temp.append(float(sheet.cell(row=x + 2, column=i).value))
            i += 1
        Xhome.append(temp)
    if sheet.cell(row=x, column=2).value == homeTeam:      # the home team today was away then
        i = 2
        temp = []
        while i < 13:
            temp.append(float(sheet.cell(row=x + 3, column=i).value))
            i += 1
        Xhome.append(temp)
    if sheet.cell(row=x, column=1).value == awayTeam:      # the away team today was home team
        i = 2
        temp = []
        while i < 13:
            temp.append(float(sheet.cell(row=x + 2, column=i).value))
            i += 1
        Xaway.append(temp)
    if sheet.cell(row=x, column=2).value == awayTeam:      # away team
        i = 2
        temp = []
        while i < 13:
            temp.append(float(sheet.cell(row=x + 3, column=i).value))
            i += 1
        Xaway.append(temp)


avgHome = [float(sum(col))/len(col) for col in zip(*Xhome)]
avgAway = [float(sum(col))/len(col) for col in zip(*Xaway)]
finalPredictionInput = avgHome + avgAway
numpy.expand_dims(finalPredictionInput, axis=0)
predictions = model.predict(numpy.array([finalPredictionInput, ]))
notRounded = [x[0] for x in predictions]
rounded = [round(x[0]) for x in predictions]
# print(Y)
print(homeTeam, awayTeam)
print(rounded, notRounded)


Xhome = []
Xaway = []
homeTeam = "New Orleans Pelicans"
awayTeam = "Golden State Warriors"
for x in range(1, 1284, 4):
    if sheet.cell(row=x, column=1).value == homeTeam:      # home team
        i = 2
        temp = []
        while i < 13:
            temp.append(float(sheet.cell(row=x + 2, column=i).value))
            i += 1
        Xhome.append(temp)
    if sheet.cell(row=x, column=2).value == homeTeam:      # the home team today was away then
        i = 2
        temp = []
        while i < 13:
            temp.append(float(sheet.cell(row=x + 3, column=i).value))
            i += 1
        Xhome.append(temp)
    if sheet.cell(row=x, column=1).value == awayTeam:      # the away team today was home team
        i = 2
        temp = []
        while i < 13:
            temp.append(float(sheet.cell(row=x + 2, column=i).value))
            i += 1
        Xaway.append(temp)
    if sheet.cell(row=x, column=2).value == awayTeam:      # away team
        i = 2
        temp = []
        while i < 13:
            temp.append(float(sheet.cell(row=x + 3, column=i).value))
            i += 1
        Xaway.append(temp)


avgHome = [float(sum(col))/len(col) for col in zip(*Xhome)]
avgAway = [float(sum(col))/len(col) for col in zip(*Xaway)]
finalPredictionInput = avgHome + avgAway
numpy.expand_dims(finalPredictionInput, axis=0)
predictions = model.predict(numpy.array([finalPredictionInput, ]))
notRounded = [x[0] for x in predictions]
rounded = [round(x[0]) for x in predictions]
# print(Y)
print(homeTeam, awayTeam)
print(rounded, notRounded)


Xhome = []
Xaway = []
homeTeam = "Dallas Mavericks"
awayTeam = "Denver Nuggets"
for x in range(1, 1284, 4):
    if sheet.cell(row=x, column=1).value == homeTeam:      # home team
        i = 2
        temp = []
        while i < 13:
            temp.append(float(sheet.cell(row=x + 2, column=i).value))
            i += 1
        Xhome.append(temp)
    if sheet.cell(row=x, column=2).value == homeTeam:      # the home team today was away then
        i = 2
        temp = []
        while i < 13:
            temp.append(float(sheet.cell(row=x + 3, column=i).value))
            i += 1
        Xhome.append(temp)
    if sheet.cell(row=x, column=1).value == awayTeam:      # the away team today was home team
        i = 2
        temp = []
        while i < 13:
            temp.append(float(sheet.cell(row=x + 2, column=i).value))
            i += 1
        Xaway.append(temp)
    if sheet.cell(row=x, column=2).value == awayTeam:      # away team
        i = 2
        temp = []
        while i < 13:
            temp.append(float(sheet.cell(row=x + 3, column=i).value))
            i += 1
        Xaway.append(temp)


avgHome = [float(sum(col))/len(col) for col in zip(*Xhome)]
avgAway = [float(sum(col))/len(col) for col in zip(*Xaway)]
finalPredictionInput = avgHome + avgAway
numpy.expand_dims(finalPredictionInput, axis=0)
predictions = model.predict(numpy.array([finalPredictionInput, ]))
notRounded = [x[0] for x in predictions]
rounded = [round(x[0]) for x in predictions]
# print(Y)
print(homeTeam, awayTeam)
print(rounded, notRounded)


Xhome = []
Xaway = []
homeTeam = "San Antonio Spurs"
awayTeam = "Detroit Pistons"
for x in range(1, 1284, 4):
    if sheet.cell(row=x, column=1).value == homeTeam:      # home team
        i = 2
        temp = []
        while i < 13:
            temp.append(float(sheet.cell(row=x + 2, column=i).value))
            i += 1
        Xhome.append(temp)
    if sheet.cell(row=x, column=2).value == homeTeam:      # the home team today was away then
        i = 2
        temp = []
        while i < 13:
            temp.append(float(sheet.cell(row=x + 3, column=i).value))
            i += 1
        Xhome.append(temp)
    if sheet.cell(row=x, column=1).value == awayTeam:      # the away team today was home team
        i = 2
        temp = []
        while i < 13:
            temp.append(float(sheet.cell(row=x + 2, column=i).value))
            i += 1
        Xaway.append(temp)
    if sheet.cell(row=x, column=2).value == awayTeam:      # away team
        i = 2
        temp = []
        while i < 13:
            temp.append(float(sheet.cell(row=x + 3, column=i).value))
            i += 1
        Xaway.append(temp)


avgHome = [float(sum(col))/len(col) for col in zip(*Xhome)]
avgAway = [float(sum(col))/len(col) for col in zip(*Xaway)]
finalPredictionInput = avgHome + avgAway
numpy.expand_dims(finalPredictionInput, axis=0)
predictions = model.predict(numpy.array([finalPredictionInput, ]))
notRounded = [x[0] for x in predictions]
rounded = [round(x[0]) for x in predictions]
# print(Y)
print(homeTeam, awayTeam)
print(rounded, notRounded)


Xhome = []
Xaway = []
homeTeam = "Utah Jazz"
awayTeam = "Washington Wizards"
for x in range(1, 1284, 4):
    if sheet.cell(row=x, column=1).value == homeTeam:      # home team
        i = 2
        temp = []
        while i < 13:
            temp.append(float(sheet.cell(row=x + 2, column=i).value))
            i += 1
        Xhome.append(temp)
    if sheet.cell(row=x, column=2).value == homeTeam:      # the home team today was away then
        i = 2
        temp = []
        while i < 13:
            temp.append(float(sheet.cell(row=x + 3, column=i).value))
            i += 1
        Xhome.append(temp)
    if sheet.cell(row=x, column=1).value == awayTeam:      # the away team today was home team
        i = 2
        temp = []
        while i < 13:
            temp.append(float(sheet.cell(row=x + 2, column=i).value))
            i += 1
        Xaway.append(temp)
    if sheet.cell(row=x, column=2).value == awayTeam:      # away team
        i = 2
        temp = []
        while i < 13:
            temp.append(float(sheet.cell(row=x + 3, column=i).value))
            i += 1
        Xaway.append(temp)


avgHome = [float(sum(col))/len(col) for col in zip(*Xhome)]
avgAway = [float(sum(col))/len(col) for col in zip(*Xaway)]
finalPredictionInput = avgHome + avgAway
numpy.expand_dims(finalPredictionInput, axis=0)
predictions = model.predict(numpy.array([finalPredictionInput, ]))
notRounded = [x[0] for x in predictions]
rounded = [round(x[0]) for x in predictions]
# print(Y)
print(homeTeam, awayTeam)
print(rounded, notRounded)
